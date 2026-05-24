
from flask import Flask, render_template, request, send_file, redirect, url_for, flash
from pathlib import Path
from datetime import datetime, date
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel
import unicodedata, re, uuid, os

app = Flask(__name__)
app.secret_key = "kardex-dashboard"

BASE = Path(__file__).resolve().parent
UPLOADS = BASE / "uploads"
OUTPUTS = BASE / "outputs"
UPLOADS.mkdir(exist_ok=True)
OUTPUTS.mkdir(exist_ok=True)

PURPLE = PatternFill(start_color="D9B3FF", end_color="D9B3FF", fill_type="solid")
HEADER = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")

def norm(v):
    if v is None:
        return ""
    t = str(v).strip().upper()
    t = unicodedata.normalize("NFKD", t)
    t = "".join(c for c in t if not unicodedata.combining(c))
    t = re.sub(r"[^A-Z0-9 ]+", " ", t)
    return re.sub(r"\s+", " ", t).strip()

def norm_name(v):
    return " ".join(sorted(norm(v).split()))

def clean_uni(v):
    return re.sub(r"\D", "", str(v or "").replace(".0", "")).lstrip("0")

def is_apto(v):
    return norm(v) == "APTO"

def as_dt(v):
    if v in [None, ""]:
        return None
    if isinstance(v, datetime):
        return v
    if isinstance(v, date):
        return datetime(v.year, v.month, v.day)
    if isinstance(v, (int, float)) and 30000 <= float(v) <= 60000:
        try:
            return from_excel(v)
        except Exception:
            pass
    for fmt in ("%d/%m/%Y %H:%M:%S","%d/%m/%Y %H:%M","%d/%m/%Y","%Y-%m-%d %H:%M:%S","%Y-%m-%d"):
        try:
            return datetime.strptime(str(v).split(".")[0].strip(), fmt)
        except Exception:
            pass
    return None

def course_key(v):
    c = norm(v)
    if c == "PLD" or "PREVENCION DE LAVADO DE DINERO" in c: return "PLD"
    if "PROTECCION DE DATOS PERSONALES" in c: return "PROTECCION DE DATOS PERSONALES"
    if "CONOCE EL MODULO DE APRENDIZAJE" in c and "USUARIO FINAL" in c: return "WORKDAY MOD APRENDIZAJE USUARIO FINAL"
    if "CONOCE EL MODULO DE APRENDIZAJE" in c and "COORDINADOR DE FORMACION" in c: return "WORKDAY MOD APRENDIZAJE COORD FORMACION"
    if "WORKDAY" in c and "APRENDIZAJE" in c and "USUARIO FINAL" in c: return "WORKDAY MOD APRENDIZAJE USUARIO FINAL"
    if "WORKDAY" in c and "APRENDIZAJE" in c and ("COORD FORMACION" in c or "COORDINADOR DE FORMACION" in c): return "WORKDAY MOD APRENDIZAJE COORD FORMACION"
    if "ORACLE SIM ROC" in c or ("ORACLE SIM" in c and "ROC" in c): return "ORACLE SIM ROC"
    if c in ["ORACLE SIM", "SISTEMA ORACLE SIM"]: return "ORACLE SIM"
    if "SICAD" in c: return "SICAD"
    if c in ["TECNICAS DE VENTA", "TECNICAS DE VENTAS"]: return "TECNICAS DE VENTAS"
    if c in ["TECNICAS DE VENTA BASICO", "TECNICAS DE VENTAS BASICO"]: return "TECNICAS DE VENTAS BASICO"
    aliases = {
        "NET PAY": "NETPAY", "AHORROPAGOS": "AHORRO PAGOS", "AHORROCEL": "AHORRO PAGOS",
        "CUENTA MONEDERO SINTETIZADA A": "CUENTA MONEDERO",
        "CUENTA MONEDERO SINTETIZADA STAFF": "CUENTA MONEDERO",
        "EL PODER DE LA MARCA PROPIA 2026": "MARCA PROPIA",
        "SERVICIO Y VENTA DE 2 ESFUERZO": "SERVICIO Y VENTAS DE 2 ESFUERZO"
    }
    return aliases.get(c, c)

def read_report(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    ws.reset_dimensions()
    header_row, headers = None, {}
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=60, values_only=True), 1):
        vals = [norm(x) for x in row]
        if "UNI" in vals and "CURSO" in vals:
            header_row = i
            headers = {norm(v): idx for idx, v in enumerate(row) if v is not None}
            break
    if header_row is None:
        raise ValueError("No encontré encabezados con UNI y CURSO en el reporte.")
    req = ["UNI","NOMBRE","CURSO","FECHA DE FINALIZACION","PUNTUACION DE CURSO","EVALUACION DE CURSO"]
    miss = [x for x in req if x not in headers]
    if miss:
        raise ValueError("Faltan columnas: " + ", ".join(miss))

    records, ignored, valid, courses = {}, 0, 0, set()
    for row in ws.iter_rows(min_row=header_row+1, values_only=True):
        try:
            u = clean_uni(row[headers["UNI"]])
            name = row[headers["NOMBRE"]]
            cur = row[headers["CURSO"]]
            dt = as_dt(row[headers["FECHA DE FINALIZACION"]])
            score = float(row[headers["PUNTUACION DE CURSO"]])
            eva = row[headers["EVALUACION DE CURSO"]]
        except Exception:
            continue
        if not u or not name or not cur:
            continue
        ck = course_key(cur)
        courses.add(ck)
        if score < 80 or not is_apto(eva) or dt is None or dt.year == 1900:
            ignored += 1
            continue
        key = (u, norm_name(name), ck)
        valid += 1
        if key not in records or dt > records[key]["dt"]:
            records[key] = {"dt": datetime(dt.year, dt.month, dt.day), "fecha": datetime(dt.year, dt.month, dt.day), "score": int(score) if score.is_integer() else score, "curso": str(cur)}
    return records, valid, ignored, courses

def detect_blocks(ws):
    blocks = []
    for col in range(1, ws.max_column + 1):
        if norm(ws.cell(5, col).value) == "FECHA DE EJECUCION" and norm(ws.cell(5, col+1).value) == "CALIFICACION":
            name = ""
            for k in range(col, max(0, col-10), -1):
                v = ws.cell(4, k).value
                if v and norm(v) not in ["FECHA DE EJECUCION","CALIFICACION","FECHA DE PROGRAMACION"]:
                    name = str(v).strip()
                    break
            blocks.append({"name": name, "key": course_key(name), "fecha": col, "cal": col+1, "prog": col-1})
    return blocks

def fix_dates(ws):
    converted, formatted = 0, 0
    fecha_cols = [c for c in range(1, ws.max_column+1) if norm(ws.cell(5,c).value) in ["FECHA DE EJECUCION","FECHA DE PROGRAMACION","FECHA DE FINALIZACION"]]
    for col in fecha_cols:
        for r in range(6, ws.max_row+1):
            cell = ws.cell(r, col)
            v = cell.value
            if isinstance(v, (int,float)) and 30000 <= float(v) <= 60000:
                dt = from_excel(v)
                cell.value = datetime(dt.year, dt.month, dt.day)
                cell.number_format = "DD/MM/YYYY"
                converted += 1
            elif isinstance(v, datetime):
                cell.value = datetime(v.year, v.month, v.day)
                cell.number_format = "DD/MM/YYYY"
                formatted += 1
    return converted, formatted

def clean_tecnicas_na(ws, blocks):
    cleaned = 0
    for b in blocks:
        if b["key"] != "TECNICAS DE VENTAS":
            continue
        allowed = set()
        for r in range(6, ws.max_row+1):
            f = ws.cell(r, b["prog"]).value
            if isinstance(f, str) and f.startswith("="):
                for ref in re.findall(r"\$?[A-Z]{1,3}\$?\d+", f):
                    ref = ref.replace("$","")
                    if not ref.startswith("H") and not ref.startswith("E"):
                        try: allowed.add(norm(ws[ref].value))
                        except Exception: pass
                break
        for r in range(6, ws.max_row+1):
            clear = norm(ws.cell(r, b["prog"]).value) == "NA" or (allowed and norm(ws.cell(r,8).value) not in allowed)
            if clear:
                f, c = ws.cell(r,b["fecha"]), ws.cell(r,b["cal"])
                if f.value not in [None,""] or c.value not in [None,""]:
                    if not (isinstance(f.value,str) and f.value.startswith("=")): f.value = None
                    if not (isinstance(c.value,str) and c.value.startswith("=")): c.value = None
                    cleaned += 1
    return cleaned

def process(kardex_path, report_path, output_path):
    records, valid, ignored, report_courses = read_report(report_path)
    wb = load_workbook(kardex_path, keep_links=False)
    ws = wb["Incorporacion"] if "Incorporacion" in wb.sheetnames else wb[wb.sheetnames[0]]
    blocks = detect_blocks(ws)
    kardex_courses = {b["key"] for b in blocks}
    no_match = sorted(report_courses - kardex_courses)
    people = {}
    for r in range(6, ws.max_row+1):
        u, name = clean_uni(ws.cell(r,6).value), norm_name(ws.cell(r,7).value)
        if u and name: people[(u,name)] = r

    changes, fechas, califs, formulas = [], 0, 0, 0
    for b in blocks:
        for (u,name,ck), rec in records.items():
            if ck != b["key"]: continue
            r = people.get((u,name))
            if not r: continue
            f, c = ws.cell(r,b["fecha"]), ws.cell(r,b["cal"])
            changed = False
            if isinstance(f.value,str) and f.value.startswith("="):
                formulas += 1
            else:
                old = as_dt(f.value)
                if f.value in [None,""] or old is None or rec["dt"] > old:
                    f.value, f.number_format, f.fill = rec["fecha"], "DD/MM/YYYY", PURPLE
                    fechas += 1; changed = True
            if isinstance(c.value,str) and c.value.startswith("="):
                formulas += 1
            else:
                if c.value != rec["score"]:
                    c.value, c.number_format, c.fill = rec["score"], "0", PURPLE
                    califs += 1; changed = True
            if changed:
                changes.append([u, ws.cell(r,7).value, rec["curso"], b["name"], rec["fecha"], rec["score"]])

    tecnicas = clean_tecnicas_na(ws, blocks)
    seriales, formateadas = fix_dates(ws)

    for sh in ["Dashboard_Actualizacion","Detalle_Cambios","Cursos_Sin_Coincidencia"]:
        if sh in wb.sheetnames: del wb[sh]

    dash = wb.create_sheet("Dashboard_Actualizacion", 0)
    rows = [
        ["KARDEX PLAZA MICHOACAN",""],["Concepto","Cantidad"],
        ["Actualizaciones realizadas", len(changes)],["Fechas cargadas/actualizadas", fechas],
        ["Calificaciones cargadas/actualizadas", califs],["Registros válidos únicos", len(records)],
        ["Registros aptos >=80 leídos", valid],["Ignorados por regla", ignored],
        ["Cursos detectados en reporte", len(report_courses)],["Cursos detectados en Kardex", len(blocks)],
        ["Cursos sin coincidencia", len(no_match)],["Fórmulas respetadas", formulas],
        ["Técnicas de Ventas limpiadas por NA", tecnicas],["Seriales convertidos a fecha", seriales],
        ["Fechas formateadas sin hora", formateadas],
    ]
    for i,row in enumerate(rows,1):
        for j,val in enumerate(row,1):
            cell = dash.cell(i,j,val)
            if i in [1,2]:
                cell.font = Font(bold=True, color="FFFFFF"); cell.fill = HEADER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    dash.column_dimensions["A"].width = 45; dash.column_dimensions["B"].width = 22

    det = wb.create_sheet("Detalle_Cambios")
    det.append(["UNI","Nombre","Curso Reporte","Curso Kardex","Fecha","Calificación"])
    for cell in det[1]:
        cell.font = Font(bold=True, color="FFFFFF"); cell.fill = HEADER
    for row in changes: det.append(row)
    for col,width in enumerate([14,40,65,55,22,18],1): det.column_dimensions[get_column_letter(col)].width = width

    nm = wb.create_sheet("Cursos_Sin_Coincidencia")
    nm.append(["Curso detectado en reporte sin coincidencia en Kardex"])
    nm["A1"].font = Font(bold=True, color="FFFFFF"); nm["A1"].fill = HEADER
    for x in no_match: nm.append([x])
    nm.column_dimensions["A"].width = 80

    wb.save(output_path)
    return {
        "filename": output_path.name, "actualizaciones": len(changes), "fechas": fechas, "calificaciones": califs,
        "validos": len(records), "ignorados": ignored, "cursos_reporte": len(report_courses),
        "cursos_kardex": len(blocks), "sin_coincidencia": len(no_match), "formulas": formulas,
        "tecnicas": tecnicas, "seriales": seriales, "top_changes": changes[:15],
        "courses_without_match": no_match[:25]
    }

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/actualizar", methods=["POST"])
def actualizar():
    try:
        kardex, reporte = request.files.get("kardex"), request.files.get("reporte")
        if not kardex or not reporte:
            flash("Debes subir el Kardex y el reporte.")
            return redirect(url_for("index"))
        job = uuid.uuid4().hex[:8]
        kp, rp, op = UPLOADS/f"kardex_{job}.xlsx", UPLOADS/f"reporte_{job}.xlsx", OUTPUTS/f"KARDEX_ACTUALIZADO_{job}.xlsx"
        kardex.save(kp); reporte.save(rp)
        result = process(kp, rp, op)
        return render_template("resultado.html", result=result)
    except Exception as e:
        flash(f"Error al procesar: {str(e)}")
        return redirect(url_for("index"))

@app.route("/descargar/<filename>")
def descargar(filename):
    return send_file(OUTPUTS / filename, as_attachment=True)

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 10000)))
